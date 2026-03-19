import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

wb = Workbook()
blue = PatternFill('solid', fgColor='2F5496')
red_bg = PatternFill('solid', fgColor='FFC7CE')
light_blue = PatternFill('solid', fgColor='D6E4F0')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
title_font = Font(name='Arial', bold=True, size=16, color='2F5496')
section_font = Font(name='Arial', bold=True, size=13, color='2F5496')
body_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
red_font = Font(name='Arial', bold=True, size=11, color='C00000')
green_font = Font(name='Arial', bold=True, size=11, color='548235')
note_font = Font(name='Arial', size=9, color='666666')
border = Border(left=Side('thin', 'D9D9D9'), right=Side('thin', 'D9D9D9'),
                top=Side('thin', 'D9D9D9'), bottom=Side('thin', 'D9D9D9'))

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header(ws, row, headers, fill=blue, font=header_font):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font
        c.fill = fill

# ============ SHEET 1: EDUCATION AS BUSINESS ============
ws1 = wb.active
ws1.title = 'Education as Business'
set_widths(ws1, [45, 20, 25])

ws1.cell(row=1, column=1, value="UK Higher Education: The Business Model").font = title_font
ws1.merge_cells('A1:C1')

ws1.cell(row=3, column=1, value="HOW MUCH INTERNATIONAL STUDENTS PAY").font = section_font
for c in range(1, 4):
    ws1.cell(row=3, column=c).fill = light_blue

money = [
    ('International fee income per year', '10 billion', 'HESA 2023/24'),
    ('Share of total university income', '20%', '1 in 5 pounds'),
    ('Growth since 2016/17', '+113%', 'From 4.7B to 10B'),
    ('Domestic fee cap (frozen)', '9,535/year', 'Has not changed'),
    ('Average international MSc fee', '15,000 - 25,000', 'Per year'),
    ('NHS surcharge (student, per year)', '776', 'Paid upfront before arrival'),
    ('NHS surcharge (PSW, per year)', '1,035', 'Paid upfront'),
    ('Visa application fee', '490 - 822', 'Per application'),
    ('Student visa drop in 2024', '-17%', 'Policy uncertainty'),
]

for i, (label, value, note) in enumerate(money):
    r = 4 + i
    ws1.cell(row=r, column=1, value=label).font = body_font
    ws1.cell(row=r, column=2, value=value).font = bold_font
    ws1.cell(row=r, column=3, value=note).font = note_font
    for c in range(1, 4): ws1.cell(row=r, column=c).border = border

r = 15
ws1.cell(row=r, column=1, value="PAWAN'S ACTUAL COSTS").font = section_font
for c in range(1, 4):
    ws1.cell(row=r, column=c).fill = red_bg

costs = [
    ('MSc tuition (Aston University)', 15000, 'International rate'),
    ('Living costs (12 months, Hull)', 10000, 'Rent + food + bills'),
    ('Student visa fee', 490, ''),
    ('NHS surcharge (student, 1 year)', 776, 'IHS'),
    ('PSW visa fee', 822, ''),
    ('NHS surcharge (PSW, 2 years)', 2070, 'IHS'),
    ('PL-300 exam fee', 165, 'Microsoft certification'),
    ('TOTAL INVESTMENT IN UK', None, ''),
]

for i, (label, value, note) in enumerate(costs):
    r2 = r + 1 + i
    ws1.cell(row=r2, column=1, value=label).font = body_font if value else red_font
    if value is not None:
        ws1.cell(row=r2, column=2, value=value).font = bold_font
        ws1.cell(row=r2, column=2).number_format = '#,##0'
    else:
        ws1.cell(row=r2, column=2, value=f'=SUM(B{r+1}:B{r2-1})').font = red_font
        ws1.cell(row=r2, column=2).number_format = '#,##0'
    ws1.cell(row=r2, column=3, value=note).font = note_font
    for c in range(1, 4): ws1.cell(row=r2, column=c).border = border

r3 = r2 + 2
ws1.cell(row=r3, column=1, value="WHAT I GOT BACK").font = section_font
for c in range(1, 4):
    ws1.cell(row=r3, column=c).fill = PatternFill('solid', fgColor='C6EFCE')

returns = [
    ('MSc degree (2:1)', 'Yes', 'Non-monetary value'),
    ('Microsoft PL-300 certification', 'Yes', 'Active, verifiable'),
    ('Starting salary after MSc', '12/hr', 'Factory work, not data role'),
    ('Data role interviews', '0', 'Zero interviews in 2 years'),
    ('Expected ROI timeline', '5+ years', 'If I could stay'),
    ('Time remaining on visa', '42 days', 'Must leave UK May 2026'),
    ('NHS services used (total)', '15', 'One A&E visit + prescription'),
    ('NHS amount paid (NI + IHS)', '4,609', '307x what I used'),
]

for i, (label, value, note) in enumerate(returns):
    r4 = r3 + 1 + i
    ws1.cell(row=r4, column=1, value=label).font = body_font
    ws1.cell(row=r4, column=2, value=value).font = bold_font
    ws1.cell(row=r4, column=3, value=note).font = note_font
    for c in range(1, 4): ws1.cell(row=r4, column=c).border = border

# ============ SHEET 2: VISA SALARY TRAP ============
ws2 = wb.create_sheet('The Salary Trap')
set_widths(ws2, [45, 20, 30])

ws2.cell(row=1, column=1, value="The Salary Trap: Why Most Graduates Cannot Stay").font = title_font
ws2.merge_cells('A1:C1')

ws2.cell(row=3, column=1, value="SKILLED WORKER VISA REQUIREMENTS (2025/26)").font = section_font

visa = [
    ('Minimum salary threshold', 41700, 'Increased from 26,200 in 2024'),
    ('Minimum hourly rate', 17.13, 'Based on 48hr week'),
    ('New entrant discount threshold', 33400, 'Only for first 3 years'),
    ('Average entry-level data analyst salary', 28000, 'Glassdoor/Indeed UK'),
    ('Gap: visa threshold vs reality', '=B4-B7', 'This is why graduates leave'),
    ('% of graduate jobs below threshold', '65%', 'Estimated'),
    ('', '', ''),
    ('GRADUATE ROUTE REALITY', '', ''),
    ('Earning within 1 month of PSW start', '62%', 'GOV.UK evaluation'),
    ('Earning within 6 months', '90%', 'GOV.UK evaluation'),
    ('In skilled (graduate-level) work', '45%', 'Less than half'),
    ('In work related to their studies', '65%', ''),
    ('Working hospitality/retail/warehouse', '~35%', 'MSc holders in factories'),
    ('Content with current situation', '65%', 'Surviving, not thriving'),
]

for i, (label, value, note) in enumerate(visa):
    r = 4 + i
    ws2.cell(row=r, column=1, value=label).font = body_font if label else section_font
    if isinstance(value, str) and value.startswith('='):
        ws2.cell(row=r, column=2, value=value).font = red_font
        ws2.cell(row=r, column=2).number_format = '#,##0'
    else:
        ws2.cell(row=r, column=2, value=value).font = bold_font
        if isinstance(value, (int, float)):
            ws2.cell(row=r, column=2).number_format = '#,##0'
    ws2.cell(row=r, column=3, value=note).font = note_font
    for c in range(1, 4): ws2.cell(row=r, column=c).border = border

ws2.cell(row=20, column=1, value="TOP BARRIERS FOR GRADUATES (GOV.UK Evaluation)").font = section_font
barriers = [
    ('1. Employers unfamiliar with visa system', '', ''),
    ('2. Employers unwilling to sponsor junior roles', '', 'Most common barrier'),
    ('3. Strong job market competition', '', ''),
    ('4. Limited visa duration (2 years only)', '', ''),
    ('5. Salary threshold too high for entry-level', '', 'The core problem'),
]
for i, (label, value, note) in enumerate(barriers):
    r = 21 + i
    ws2.cell(row=r, column=1, value=label).font = body_font
    ws2.cell(row=r, column=3, value=note).font = note_font

# ============ SHEET 3: ATTAINMENT DATA ============
ws3 = wb.create_sheet('Attainment by Ethnicity')
set_widths(ws3, [35, 18, 18, 18, 18])

ws3.cell(row=1, column=1, value="A-Level Attainment by Ethnicity (2024/25)").font = title_font
write_header(ws3, 3, ['Ethnic Group', 'Total Students', 'Achieved Level 3', 'Attainment Rate %', 'Gap from Best'])

alevel = pd.read_csv('data/raw/alevel_ethnicity.csv')
latest = alevel['time_period'].max()
groups = ['White', 'Asian or Asian British', 'Black or Black British',
          'Mixed Dual background', 'Any other ethnic group']
filt = alevel[(alevel['time_period'] == latest) & (alevel['sex'] == 'All') &
              (alevel['disadvantage'] == 'All') & (alevel['sen'] == 'All') &
              (alevel['ethnicity_minor'].isin(groups))]

rates = []
for _, s in filt.iterrows():
    total = pd.to_numeric(s.get('number_of_students_potential', 0), errors='coerce')
    l3 = pd.to_numeric(s.get('number_of_students_level3', 0), errors='coerce')
    rate = (l3 / total * 100) if total > 0 else 0
    rates.append((s['ethnicity_minor'], int(total) if pd.notna(total) else 0,
                  int(l3) if pd.notna(l3) else 0, round(rate, 1)))

max_rate = max(r[3] for r in rates) if rates else 0
for i, (eth, total, l3, rate) in enumerate(rates):
    r = 4 + i
    ws3.cell(row=r, column=1, value=eth).font = body_font
    ws3.cell(row=r, column=2, value=total).font = body_font
    ws3.cell(row=r, column=2).number_format = '#,##0'
    ws3.cell(row=r, column=3, value=l3).font = body_font
    ws3.cell(row=r, column=3).number_format = '#,##0'
    ws3.cell(row=r, column=4, value=rate).font = bold_font
    gap = round(rate - max_rate, 1)
    ws3.cell(row=r, column=5, value=gap).font = red_font if gap < 0 else green_font
    for c in range(1, 6): ws3.cell(row=r, column=c).border = border

# ============ SHEET 4: PAWAN'S JOURNEY ============
ws4 = wb.create_sheet("Pawan's Journey")
set_widths(ws4, [18, 50, 18, 18])

ws4.cell(row=1, column=1, value="My UK Journey: What I Put In vs What I Got Back").font = title_font
ws4.merge_cells('A1:D1')
write_header(ws4, 3, ['Date', 'Event', 'Cost/Earned', 'Running Total'])

timeline = [
    ('Sep 2023', 'Arrived UK for MSc at Aston University', -15000),
    ('Sep 2023', 'Visa fee + NHS surcharge (student)', -1266),
    ('Sep 2023-Jun 2024', 'Living costs (9 months)', -9000),
    ('Mar 2024', 'Graduated MSc Data Analytics (2:1)', 0),
    ('Mar-Jul 2024', 'Stadium cashier + hospitality (self-studied AWS)', 3500),
    ('May 2024', 'PSW visa fee + NHS surcharge', -2892),
    ('Jun 2024', 'Relocated Birmingham to Hull (cheaper living)', -200),
    ('Jul 2024', 'Started Cranswick (food operative, 12/hr)', 0),
    ('Oct 2024', 'Promoted to Cover Team Lead', 0),
    ('Apr 2025', 'Moved to Copernus (Team Leader, ERP systems)', 0),
    ('Jul 2024-Mar 2026', 'Total earnings from employment (estimated)', 22000),
    ('Mar 2026', 'Passed Microsoft PL-300 certification', -165),
    ('Mar 2026', 'Built 10 portfolio projects on GitHub', 0),
    ('May 2026', 'PSW visa expires - must leave UK', 0),
]

running = 0
for i, (date, event, cost) in enumerate(timeline):
    r = 4 + i
    ws4.cell(row=r, column=1, value=date).font = body_font
    ws4.cell(row=r, column=2, value=event).font = body_font
    ws4.cell(row=r, column=3, value=cost).font = red_font if cost < 0 else (green_font if cost > 0 else body_font)
    ws4.cell(row=r, column=3).number_format = '#,##0'
    running += cost
    ws4.cell(row=r, column=4, value=running).font = bold_font
    ws4.cell(row=r, column=4).number_format = '#,##0'
    for c in range(1, 5): ws4.cell(row=r, column=c).border = border

r_final = 4 + len(timeline)
ws4.cell(row=r_final, column=1, value='NET POSITION').font = red_font
ws4.cell(row=r_final, column=2, value='Total cost to me after earnings').font = red_font
ws4.cell(row=r_final, column=3, value=f'=SUM(C4:C{r_final-1})').font = red_font
ws4.cell(row=r_final, column=3).number_format = '#,##0'

# ============ SHEET 5: KEY FINDINGS ============
ws5 = wb.create_sheet('Key Findings')
ws5.column_dimensions['A'].width = 70

ws5.cell(row=1, column=1, value="Key Findings: Education, Immigration, and Policy").font = title_font

findings = [
    (3, "FINDING 1: UK universities depend on international students for survival", True),
    (4, "International student fees = 10B/year. Without us, 20% of university income disappears.", False),
    (5, "This is not education policy. This is revenue extraction.", False),
    (7, "FINDING 2: The visa system is designed to extract money, not retain talent", True),
    (8, "I paid ~29,000. Skilled Worker visa requires 41,700/year. Entry data analyst pays 28,000.", False),
    (9, "The gap is 13,700/year. There is no legitimate pathway for most graduates.", False),
    (11, "FINDING 3: Immigrant-background students outperform in education", True),
    (12, "Chinese and Indian students achieve highest A-Level attainment. Black African students outperform", False),
    (13, "White British in many metrics. The UK benefits from immigrant talent — then discards it.", False),
    (15, "FINDING 4: Deprivation is the real predictor, not ethnicity", True),
    (16, "A disadvantaged White student has worse outcomes than a non-disadvantaged Black student.", False),
    (17, "Class, not race, drives the attainment gap. Policy addresses neither effectively.", False),
    (19, "FINDING 5: 10 years of policy, minimal change", True),
    (20, "Widening participation schemes cost billions. The attainment gap barely moved.", False),
    (21, "The system produces reports about the gap. It does not close the gap.", False),
    (23, "FINDING 6: The cycle is the product", True),
    (24, "Step 1: Recruit international students at 2-3x domestic fees.", False),
    (25, "Step 2: Collect visa fees and NHS surcharges upfront.", False),
    (26, "Step 3: Graduate them into a job market that will not sponsor them.", False),
    (27, "Step 4: They leave. The next cohort arrives. Repeat.", False),
    (28, "This is not a bug. This is the business model.", False),
    (30, "MY CONCLUSION", True),
    (31, "I invested 29,000 and 3 years in the UK. I earned an MSc, passed certifications,", False),
    (32, "built 10 data projects, paid taxes, and used 15 of NHS services.", False),
    (33, "The system took my money and my time. In 42 days it will ask me to leave.", False),
    (34, "The data does not lie. The policy does.", False),
]

for r, text, is_heading in findings:
    ws5.cell(row=r, column=1, value=text).font = section_font if is_heading else body_font

output = 'UK_Education_Investigation.xlsx'
wb.save(output)
print(f'Saved: {output}')
print('Sheets: Education as Business, The Salary Trap, Attainment by Ethnicity, Pawans Journey, Key Findings')
