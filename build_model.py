"""
UK Graduate Talent Pipeline Analysis
=====================================
An evidence-based assessment of the UK's international graduate talent pipeline.
Built as an updateable model — feed new data as government statistics refresh.

Author: Pawan Singh Kapkoti
Data: Department for Education, HESA, GOV.UK, ONS
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

wb = Workbook()

# Styles
BLUE = PatternFill('solid', fgColor='2F5496')
LIGHT_BLUE = PatternFill('solid', fgColor='D6E4F0')
INPUT_YELLOW = PatternFill('solid', fgColor='FFF2CC')
OUTPUT_GREEN = PatternFill('solid', fgColor='E2EFDA')
ALERT_RED = PatternFill('solid', fgColor='FFC7CE')
WHITE = PatternFill('solid', fgColor='FFFFFF')

H1 = Font(name='Arial', bold=True, size=16, color='2F5496')
H2 = Font(name='Arial', bold=True, size=13, color='2F5496')
H3 = Font(name='Arial', bold=True, size=11, color='2F5496')
HEADER = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY = Font(name='Arial', size=10)
BOLD = Font(name='Arial', bold=True, size=10)
INPUT_FONT = Font(name='Arial', bold=True, size=10, color='0000FF')
OUTPUT_FONT = Font(name='Arial', bold=True, size=10, color='006100')
ALERT_FONT = Font(name='Arial', bold=True, size=10, color='C00000')
NOTE = Font(name='Arial', size=9, color='666666')
BORDER = Border(left=Side('thin', 'D9D9D9'), right=Side('thin', 'D9D9D9'),
                top=Side('thin', 'D9D9D9'), bottom=Side('thin', 'D9D9D9'))

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_row(ws, row, col_data, fonts=None, fills=None):
    for i, val in enumerate(col_data):
        c = ws.cell(row=row, column=i+1, value=val)
        c.font = fonts[i] if fonts else BODY
        if fills and fills[i]: c.fill = fills[i]
        c.border = BORDER

def section_header(ws, row, text, cols=5):
    ws.cell(row=row, column=1, value=text).font = H2
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).fill = LIGHT_BLUE

# ============================================================
# SHEET 1: MODEL INPUTS (Updateable)
# ============================================================
ws1 = wb.active
ws1.title = 'Model Inputs'
set_widths(ws1, [45, 20, 15, 35])

ws1.cell(row=1, column=1, value="UK Graduate Talent Pipeline — Model Inputs").font = H1
ws1.merge_cells('A1:D1')
ws1.cell(row=2, column=1, value="Update BLUE cells when new data is published. All other sheets recalculate automatically.").font = NOTE

section_header(ws1, 4, "GOVERNMENT FEE STRUCTURE", 4)
write_row(ws1, 5, ['Parameter', 'Value', 'Unit', 'Source / Update Frequency'], [HEADER]*4, [BLUE]*4)

inputs = [
    ('Student visa application fee', 490, 'GBP', 'GOV.UK — updated annually'),
    ('PSW (Graduate) visa application fee', 822, 'GBP', 'GOV.UK — updated annually'),
    ('NHS surcharge rate (student, per year)', 776, 'GBP/yr', 'GOV.UK — updated annually'),
    ('NHS surcharge rate (PSW, per year)', 1035, 'GBP/yr', 'GOV.UK — updated annually'),
    ('Skilled Worker visa minimum salary', 41700, 'GBP/yr', 'GOV.UK — last changed Jul 2025'),
    ('Skilled Worker new entrant threshold', 33400, 'GBP/yr', 'GOV.UK — last changed Jul 2025'),
    ('Skilled Worker minimum hourly rate', 17.13, 'GBP/hr', 'GOV.UK — based on 48hr week'),
]

for i, (label, value, unit, source) in enumerate(inputs):
    r = 6 + i
    ws1.cell(row=r, column=1, value=label).font = BODY
    ws1.cell(row=r, column=2, value=value).font = INPUT_FONT
    ws1.cell(row=r, column=2).fill = INPUT_YELLOW
    ws1.cell(row=r, column=2).number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
    ws1.cell(row=r, column=3, value=unit).font = NOTE
    ws1.cell(row=r, column=4, value=source).font = NOTE
    for c in range(1, 5): ws1.cell(row=r, column=c).border = BORDER

r = 14
section_header(ws1, r, "UNIVERSITY FEE STRUCTURE", 4)
write_row(ws1, r+1, ['Parameter', 'Value', 'Unit', 'Source'], [HEADER]*4, [BLUE]*4)

uni_inputs = [
    ('Average international MSc fee (non-London)', 15500, 'GBP/yr', 'HESA — updated annually'),
    ('Average international MSc fee (London)', 22000, 'GBP/yr', 'HESA — updated annually'),
    ('Domestic undergraduate fee cap', 9535, 'GBP/yr', 'DfE — last changed 2025'),
    ('International fee premium (multiplier)', None, 'x', '=B{0}/B{2}'.format(r+2, r+2, r+4)),
    ('Total international fee income (sector)', 10000000000, 'GBP', 'HESA 2023/24'),
    ('International share of total fee income', 0.45, '%', 'HESA — 45p of every pound'),
]

for i, (label, value, unit, source) in enumerate(uni_inputs):
    r2 = r + 2 + i
    ws1.cell(row=r2, column=1, value=label).font = BODY
    if value is not None:
        ws1.cell(row=r2, column=2, value=value).font = INPUT_FONT
        ws1.cell(row=r2, column=2).fill = INPUT_YELLOW
    else:
        ws1.cell(row=r2, column=2, value=f'=B{r+2}/B{r+4}').font = OUTPUT_FONT
        ws1.cell(row=r2, column=2).fill = OUTPUT_GREEN
    if isinstance(value, float):
        ws1.cell(row=r2, column=2).number_format = '0.0%'
    elif isinstance(value, int) and value > 100000:
        ws1.cell(row=r2, column=2).number_format = '#,##0'
    else:
        ws1.cell(row=r2, column=2).number_format = '#,##0'
    ws1.cell(row=r2, column=3, value=unit).font = NOTE
    ws1.cell(row=r2, column=4, value=source).font = NOTE
    for c in range(1, 5): ws1.cell(row=r2, column=c).border = BORDER

r = 23
section_header(ws1, r, "LABOUR MARKET DATA", 4)
write_row(ws1, r+1, ['Parameter', 'Value', 'Unit', 'Source'], [HEADER]*4, [BLUE]*4)

labour = [
    ('Median graduate salary (all subjects)', 42000, 'GBP/yr', 'ONS ASHE 2024'),
    ('Median postgraduate salary', 47000, 'GBP/yr', 'ONS ASHE 2024'),
    ('Entry-level data analyst salary', 28000, 'GBP/yr', 'Glassdoor/Indeed avg'),
    ('Entry-level data engineer salary', 32000, 'GBP/yr', 'Glassdoor/Indeed avg'),
    ('National Living Wage (hourly)', 12.21, 'GBP/hr', 'GOV.UK — April 2025'),
    ('PSW graduates earning within 1 month', 0.62, '%', 'GOV.UK evaluation 2024'),
    ('PSW graduates in skilled employment', 0.45, '%', 'Graduate Outcomes Survey'),
    ('PSW graduates in study-related work', 0.65, '%', 'GOV.UK evaluation 2024'),
]

for i, (label, value, unit, source) in enumerate(labour):
    r2 = r + 2 + i
    ws1.cell(row=r2, column=1, value=label).font = BODY
    ws1.cell(row=r2, column=2, value=value).font = INPUT_FONT
    ws1.cell(row=r2, column=2).fill = INPUT_YELLOW
    if isinstance(value, float) and value < 1:
        ws1.cell(row=r2, column=2).number_format = '0.0%'
    elif isinstance(value, float):
        ws1.cell(row=r2, column=2).number_format = '#,##0.00'
    else:
        ws1.cell(row=r2, column=2).number_format = '#,##0'
    ws1.cell(row=r2, column=3, value=unit).font = NOTE
    ws1.cell(row=r2, column=4, value=source).font = NOTE
    for c in range(1, 5): ws1.cell(row=r2, column=c).border = BORDER

# ============================================================
# SHEET 2: CALCULATED ANALYSIS
# ============================================================
ws2 = wb.create_sheet('Pipeline Analysis')
set_widths(ws2, [45, 20, 20, 30])

ws2.cell(row=1, column=1, value="UK Graduate Talent Pipeline — Calculated Analysis").font = H1
ws2.merge_cells('A1:D1')
ws2.cell(row=2, column=1, value="All values auto-calculate from Model Inputs sheet. Green cells = formulas.").font = NOTE

section_header(ws2, 4, "COST-BENEFIT: INTERNATIONAL MSc GRADUATE", 4)
write_row(ws2, 5, ['Metric', 'Amount (GBP)', 'Calculation', 'Finding'], [HEADER]*4, [BLUE]*4)

calcs = [
    ('Total visa + NHS costs (student year)', "='Model Inputs'!B6+'Model Inputs'!B8", 'Visa + 1yr NHS', ''),
    ('Total visa + NHS costs (PSW 2 years)', "='Model Inputs'!B7+'Model Inputs'!B9*2", 'Visa + 2yr NHS', ''),
    ('Total government fees paid', '=B6+B7', 'Sum of above', ''),
    ('Tuition fee paid (non-London MSc)', "='Model Inputs'!B16", 'Direct to university', ''),
    ('Total cost to study + stay 3 years', '=B8+B9', 'Fees + tuition', ''),
    ('Estimated living costs (3 years)', 30000, '', '10k/yr average'),
    ('TOTAL INVESTMENT BY GRADUATE', '=B10+B11', '', ''),
    ('', '', '', ''),
    ('Revenue to UK per international MSc grad', '=B12', '', 'What the UK receives'),
    ('Skilled Worker salary threshold', "='Model Inputs'!B10", '', 'What they need to earn to stay'),
    ('Typical entry salary achieved', "='Model Inputs'!B27", '', 'What they actually earn'),
    ('SALARY GAP (threshold - reality)', '=B15-B16', '', 'This gap forces graduates out'),
    ('Gap as % of entry salary', '=B17/B16', '', ''),
]

for i, (label, value, calc, finding) in enumerate(calcs):
    r = 6 + i
    ws2.cell(row=r, column=1, value=label).font = BOLD if label.startswith('TOTAL') or label.startswith('SALARY') else BODY
    if isinstance(value, str) and value.startswith('='):
        ws2.cell(row=r, column=2, value=value).font = OUTPUT_FONT
        ws2.cell(row=r, column=2).fill = OUTPUT_GREEN
    elif isinstance(value, (int, float)):
        ws2.cell(row=r, column=2, value=value).font = INPUT_FONT
        ws2.cell(row=r, column=2).fill = INPUT_YELLOW
    ws2.cell(row=r, column=2).number_format = '0.0%' if 'as %' in label else '#,##0'
    ws2.cell(row=r, column=3, value=calc).font = NOTE
    ws2.cell(row=r, column=4, value=finding).font = NOTE
    for c in range(1, 5): ws2.cell(row=r, column=c).border = BORDER

r = 21
section_header(ws2, r, "TALENT RETENTION RATE", 4)
write_row(ws2, r+1, ['Metric', 'Value', 'Calculation', 'Implication'], [HEADER]*4, [BLUE]*4)

retention = [
    ('International students graduating per year (est)', 200000, '', 'HESA estimate'),
    ('PSW visa uptake rate', 0.60, '', 'Not all apply'),
    ('Graduates entering skilled employment', "='Model Inputs'!B31", '', ''),
    ('Graduates who could meet salary threshold', 0.15, '', 'Estimated 15% earn 41,700+'),
    ('Skilled Worker visa conversions (est)', '=B{0}*B{1}'.format(r+2, r+5), '', ''),
    ('TALENT RETAINED (%)', '=B{0}/B{1}'.format(r+6, r+2), '', 'Of all graduates'),
    ('TALENT LOST (%)', '=1-B{0}'.format(r+7), '', 'Trained then exported'),
    ('Cost to train talent that leaves', '=B{0}*B12'.format(r+8), '', 'Economic waste'),
]

for i, (label, value, calc, impl) in enumerate(retention):
    r2 = r + 2 + i
    ws2.cell(row=r2, column=1, value=label).font = BOLD if label.startswith('TALENT') or label.startswith('Cost') else BODY
    if isinstance(value, str) and value.startswith('='):
        ws2.cell(row=r2, column=2, value=value).font = OUTPUT_FONT
        ws2.cell(row=r2, column=2).fill = OUTPUT_GREEN
    elif isinstance(value, float) and value < 1:
        ws2.cell(row=r2, column=2, value=value).font = INPUT_FONT
        ws2.cell(row=r2, column=2).fill = INPUT_YELLOW
        ws2.cell(row=r2, column=2).number_format = '0.0%'
    elif isinstance(value, (int, float)):
        ws2.cell(row=r2, column=2, value=value).font = INPUT_FONT
        ws2.cell(row=r2, column=2).fill = INPUT_YELLOW
        ws2.cell(row=r2, column=2).number_format = '#,##0'
    ws2.cell(row=r2, column=3, value=calc).font = NOTE
    ws2.cell(row=r2, column=4, value=impl).font = NOTE
    for c in range(1, 5): ws2.cell(row=r2, column=c).border = BORDER

# ============================================================
# SHEET 3: ATTAINMENT DATA
# ============================================================
ws3 = wb.create_sheet('Attainment Evidence')
set_widths(ws3, [35, 18, 18, 18, 18])

ws3.cell(row=1, column=1, value="A-Level Attainment by Ethnicity — Government Data").font = H1
ws3.merge_cells('A1:E1')
ws3.cell(row=2, column=1, value="Source: Department for Education, explore-education-statistics.service.gov.uk").font = NOTE

write_row(ws3, 4, ['Ethnic Group', 'Total Students', 'Level 3 Achieved', 'Rate (%)', 'Gap from Best (pp)'],
          [HEADER]*5, [BLUE]*5)

alevel = pd.read_csv('data/raw/alevel_ethnicity.csv')
latest = alevel['time_period'].max()
groups = ['White', 'Asian or Asian British', 'Black or Black British',
          'Mixed Dual background', 'Any other ethnic group',
          'Asian or Asian British - Indian', 'Asian or Asian British - Pakistani',
          'Asian or Asian British - Bangladeshi', 'Asian or Asian British - Chinese',
          'Black or Black British - Black African', 'Black or Black British - Black Caribbean']
available = [g for g in groups if g in alevel['ethnicity_minor'].unique()]

filt = alevel[(alevel['time_period'] == latest) & (alevel['sex'] == 'All') &
              (alevel['disadvantage'] == 'All') & (alevel['sen'] == 'All') &
              (alevel['ethnicity_minor'].isin(available))]

rates = []
for _, s in filt.iterrows():
    total = pd.to_numeric(s.get('number_of_students_potential', 0), errors='coerce')
    l3 = pd.to_numeric(s.get('number_of_students_level3', 0), errors='coerce')
    rate = (l3 / total * 100) if pd.notna(total) and total > 0 else 0
    rates.append((s['ethnicity_minor'], int(total) if pd.notna(total) else 0,
                  int(l3) if pd.notna(l3) else 0, round(rate, 1)))

rates.sort(key=lambda x: x[3], reverse=True)
max_rate = rates[0][3] if rates else 0

for i, (eth, total, l3, rate) in enumerate(rates):
    r = 5 + i
    ws3.cell(row=r, column=1, value=eth).font = BODY
    ws3.cell(row=r, column=2, value=total).font = BODY
    ws3.cell(row=r, column=2).number_format = '#,##0'
    ws3.cell(row=r, column=3, value=l3).font = BODY
    ws3.cell(row=r, column=3).number_format = '#,##0'
    ws3.cell(row=r, column=4, value=rate).font = BOLD
    gap = round(rate - max_rate, 1)
    ws3.cell(row=r, column=5, value=gap).font = ALERT_FONT if gap < -5 else (OUTPUT_FONT if gap >= 0 else BODY)
    for c in range(1, 6): ws3.cell(row=r, column=c).border = BORDER

r_note = 5 + len(rates) + 2
ws3.cell(row=r_note, column=1, value="ANALYST NOTE:").font = H3
ws3.cell(row=r_note+1, column=1, value="Communities with strong immigration ties (Indian, Chinese, Black African) consistently").font = BODY
ws3.cell(row=r_note+2, column=1, value="achieve attainment rates at or above the national average. This contradicts the narrative").font = BODY
ws3.cell(row=r_note+3, column=1, value="that immigration lowers educational standards. The data shows the opposite.").font = BODY

# ============================================================
# SHEET 4: SCENARIO MODELLING
# ============================================================
ws4 = wb.create_sheet('Scenario Model')
set_widths(ws4, [40, 20, 20, 20])

ws4.cell(row=1, column=1, value="What If? — Policy Scenario Modelling").font = H1
ws4.merge_cells('A1:D1')
ws4.cell(row=2, column=1, value="Change BLUE input cells to model different policy scenarios.").font = NOTE

section_header(ws4, 4, "SCENARIO: LOWER SALARY THRESHOLD", 4)
write_row(ws4, 5, ['Parameter', 'Current', 'Proposed', 'Impact'], [HEADER]*4, [BLUE]*4)

scenarios = [
    ('Salary threshold', "='Model Inputs'!B10", 30000, ''),
    ('Entry-level data analyst salary', "='Model Inputs'!B27", "='Model Inputs'!B27", ''),
    ('Gap (threshold - salary)', '=B6-B7', '=C6-C7', ''),
    ('Can a data analyst qualify?', '=IF(B8>0,"NO","YES")', '=IF(C8>0,"NO","YES")', ''),
    ('Estimated additional graduates retained', '', 25000, 'Per year estimate'),
    ('Additional tax revenue (20% rate)', '', '=C10*C6*0.2', 'Income tax alone'),
    ('Additional NI revenue', '', '=C10*C6*0.12', 'Employee NI'),
    ('Additional consumer spending', '', '=C10*C6*0.5', '50% spent in UK economy'),
    ('TOTAL ECONOMIC GAIN', '', '=C11+C12+C13', ''),
]

for i, (label, current, proposed, impact) in enumerate(scenarios):
    r = 6 + i
    ws4.cell(row=r, column=1, value=label).font = BOLD if label.startswith('TOTAL') or label.startswith('Can') else BODY
    if isinstance(current, str) and current.startswith('='):
        ws4.cell(row=r, column=2, value=current).font = OUTPUT_FONT
        ws4.cell(row=r, column=2).fill = OUTPUT_GREEN
    else:
        ws4.cell(row=r, column=2, value=current).font = BODY
    if isinstance(proposed, str) and proposed.startswith('='):
        ws4.cell(row=r, column=3, value=proposed).font = OUTPUT_FONT
        ws4.cell(row=r, column=3).fill = OUTPUT_GREEN
    elif isinstance(proposed, (int, float)):
        ws4.cell(row=r, column=3, value=proposed).font = INPUT_FONT
        ws4.cell(row=r, column=3).fill = INPUT_YELLOW
    ws4.cell(row=r, column=4, value=impact).font = NOTE
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = BORDER
        if isinstance(ws4.cell(row=r, column=c).value, (int, float)) or (isinstance(ws4.cell(row=r, column=c).value, str) and ws4.cell(row=r, column=c).value.startswith('=')):
            ws4.cell(row=r, column=c).number_format = '#,##0'

r = 17
section_header(ws4, r, "SCENARIO: GRADUATE RETENTION PROGRAMME", 4)
ws4.cell(row=r+1, column=1, value="If the UK retained 50% of international MSc graduates for 5 years:").font = BODY
ws4.cell(row=r+2, column=1, value="Graduates retained per year").font = BODY
ws4.cell(row=r+2, column=2, value=100000).font = INPUT_FONT
ws4.cell(row=r+2, column=2).fill = INPUT_YELLOW
ws4.cell(row=r+3, column=1, value="Average salary earned").font = BODY
ws4.cell(row=r+3, column=2, value=35000).font = INPUT_FONT
ws4.cell(row=r+3, column=2).fill = INPUT_YELLOW
ws4.cell(row=r+4, column=1, value="Annual income tax generated").font = BODY
ws4.cell(row=r+4, column=2, value=f'=B{r+2}*B{r+3}*0.2').font = OUTPUT_FONT
ws4.cell(row=r+4, column=2).fill = OUTPUT_GREEN
ws4.cell(row=r+4, column=2).number_format = '#,##0'
ws4.cell(row=r+5, column=1, value="Annual NI generated").font = BODY
ws4.cell(row=r+5, column=2, value=f'=B{r+2}*B{r+3}*0.12').font = OUTPUT_FONT
ws4.cell(row=r+5, column=2).fill = OUTPUT_GREEN
ws4.cell(row=r+5, column=2).number_format = '#,##0'
ws4.cell(row=r+6, column=1, value="TOTAL TAX REVENUE OVER 5 YEARS").font = BOLD
ws4.cell(row=r+6, column=2, value=f'=(B{r+4}+B{r+5})*5').font = OUTPUT_FONT
ws4.cell(row=r+6, column=2).fill = OUTPUT_GREEN
ws4.cell(row=r+6, column=2).number_format = '#,##0'

# ============================================================
# SHEET 5: FINDINGS & RECOMMENDATIONS
# ============================================================
ws5 = wb.create_sheet('Findings')
ws5.column_dimensions['A'].width = 8
ws5.column_dimensions['B'].width = 65

ws5.cell(row=1, column=1, value="UK Graduate Talent Pipeline — Findings & Recommendations").font = H1
ws5.merge_cells('A1:B1')

findings = [
    (3, '#', 'Finding', True),
    (4, '1', 'The UK higher education sector generates 10B/year from international students — 20% of total income. This is an export industry, not a charity.', False),
    (5, '2', 'The Skilled Worker visa salary threshold (41,700) exceeds the entry-level salary for most graduate professions (25-32k). This creates a structural barrier that prevents graduates from transitioning to sponsored employment.', False),
    (6, '3', 'Only 45% of PSW graduates secure skilled employment. 35% work in hospitality, retail, or warehouse roles despite holding postgraduate qualifications.', False),
    (7, '4', 'Government data shows that students from immigrant-background communities (Indian, Chinese, Black African) consistently achieve attainment rates at or above the national average at A-Level. Immigration strengthens educational outcomes.', False),
    (8, '5', 'Deprivation is a stronger predictor of attainment than ethnicity. The attainment gap between disadvantaged and non-disadvantaged students within the same ethnic group exceeds the gap between ethnic groups.', False),
    (9, '6', 'The attainment gap has not significantly narrowed over 10 years of widening participation policy. Current interventions are insufficient.', False),
    (10, '7', 'The current pipeline model extracts maximum revenue from international students (tuition + visa fees + NHS surcharge) then structurally prevents most from contributing to the UK labour market. This represents an economic inefficiency.', False),
    (12, '', '', True),
    (13, '#', 'Recommendation', True),
    (14, 'R1', 'Introduce a post-study salary threshold aligned with actual graduate salaries (28-30k) rather than the general Skilled Worker threshold (41,700). This retains skilled graduates who are ready to contribute.', False),
    (15, 'R2', 'Create a Graduate Retention Pathway allowing graduates in shortage occupations to transition to Skilled Worker visas at a reduced salary threshold for the first 3 years.', False),
    (16, 'R3', 'Publish annual data on PSW graduate employment outcomes by subject, university, and salary band. Current reporting lacks the granularity needed for evidence-based policy.', False),
    (17, 'R4', 'Model the fiscal impact of graduate retention vs departure. This analysis estimates that retaining 100,000 graduates at 35k average salary would generate 5.6B in tax and NI over 5 years.', False),
    (18, 'R5', 'Address deprivation as the primary driver of the attainment gap, rather than ethnicity. Current policy focuses on the wrong variable.', False),
]

for r, col1, col2, is_header in findings:
    ws5.cell(row=r, column=1, value=col1).font = HEADER if is_header else BOLD
    ws5.cell(row=r, column=2, value=col2).font = H3 if is_header else BODY
    if is_header:
        ws5.cell(row=r, column=1).fill = BLUE
        ws5.cell(row=r, column=2).fill = BLUE
    ws5.cell(row=r, column=2).alignment = Alignment(wrap_text=True)

ws5.cell(row=20, column=1, value="").font = BODY
ws5.cell(row=21, column=1, value="DATA").font = H3
ws5.cell(row=22, column=2, value="Department for Education — explore-education-statistics.service.gov.uk").font = NOTE
ws5.cell(row=23, column=2, value="HESA Higher Education Student Statistics — hesa.ac.uk").font = NOTE
ws5.cell(row=24, column=2, value="GOV.UK Graduate Route Evaluation — gov.uk/government/publications/graduate-route-evaluation").font = NOTE
ws5.cell(row=25, column=2, value="ONS Annual Survey of Hours and Earnings — ons.gov.uk").font = NOTE

ws5.cell(row=27, column=1, value="AUTHOR").font = H3
ws5.cell(row=28, column=2, value="Pawan Singh Kapkoti — MSc Data Analytics, Aston University").font = BODY
ws5.cell(row=29, column=2, value="Microsoft Certified: Power BI Data Analyst Associate (PL-300)").font = BODY
ws5.cell(row=30, column=2, value="github.com/Pawansingh3889 | linkedin.com/in/pawan-singh-kapkoti-100176347").font = NOTE

# ============================================================
# SHEET 6: CASE STUDY — SINGLE GRADUATE PROFILE
# ============================================================
ws6 = wb.create_sheet('Case Study - Graduate Profile')
set_widths(ws6, [35, 20, 20, 30])

ws6.cell(row=1, column=1, value="Case Study: Single Graduate Pipeline Journey").font = H1
ws6.merge_cells('A1:D1')
ws6.cell(row=2, column=1, value="One data point in the pipeline. Update blue cells to model any graduate.").font = NOTE

section_header(ws6, 4, "GRADUATE PROFILE (Updateable)", 4)
write_row(ws6, 5, ['Field', 'Value', 'Unit', 'Notes'], [HEADER]*4, [BLUE]*4)

profile = [
    ('Name', 'Pawan Singh Kapkoti', '', ''),
    ('Nationality', 'Indian', '', ''),
    ('Qualification', 'MSc Data Analytics', '', 'Aston University, 2:1'),
    ('Graduation date', 'March 2024', '', ''),
    ('Visa type', 'PSW (Graduate Route)', '', '2 year duration'),
    ('Visa expiry', 'May 2026', '', ''),
    ('Days remaining (as of analysis)', 42, 'days', 'Auto-update when reviewing'),
    ('Certifications earned in UK', 'PL-300, Google Data Analytics', '', 'Self-funded, self-studied'),
    ('Portfolio projects built', 10, '', 'All public on GitHub'),
    ('Target role', 'Data Analyst / Data Engineer', '', 'SOC 3544 / 3133'),
    ('Target salary', 28000, 'GBP/yr', 'Entry level data analyst'),
]

for i, (field, value, unit, notes) in enumerate(profile):
    r = 6 + i
    ws6.cell(row=r, column=1, value=field).font = BODY
    ws6.cell(row=r, column=2, value=value).font = INPUT_FONT
    ws6.cell(row=r, column=2).fill = INPUT_YELLOW
    if isinstance(value, int): ws6.cell(row=r, column=2).number_format = '#,##0'
    ws6.cell(row=r, column=3, value=unit).font = NOTE
    ws6.cell(row=r, column=4, value=notes).font = NOTE
    for c in range(1, 5): ws6.cell(row=r, column=c).border = BORDER

r = 19
section_header(ws6, r, "INVESTMENT INTO UK ECONOMY", 4)
write_row(ws6, r+1, ['Item', 'Amount (GBP)', 'Recipient', 'Type'], [HEADER]*4, [BLUE]*4)

investment = [
    ('MSc tuition fee', 15000, 'Aston University', 'Education'),
    ('Student visa fee', 490, 'Home Office', 'Government fee'),
    ('NHS surcharge (student, 1 year)', 776, 'NHS / Treasury', 'Healthcare access'),
    ('PSW visa fee', 822, 'Home Office', 'Government fee'),
    ('NHS surcharge (PSW, 2 years)', 2070, 'NHS / Treasury', 'Healthcare access'),
    ('Living costs (rent, food, bills - 30 months)', 25000, 'UK landlords, businesses', 'Consumer spending'),
    ('PL-300 exam fee', 165, 'Microsoft (UK operation)', 'Certification'),
    ('Income tax paid (20 months employment)', 3200, 'HMRC', 'Tax revenue'),
    ('National Insurance paid (20 months)', 2400, 'HMRC', 'Tax revenue'),
    ('Council tax paid (20 months)', 1600, 'Hull City Council', 'Local government'),
    ('TOTAL INVESTED INTO UK', None, '', ''),
]

for i, (item, amount, recipient, category) in enumerate(investment):
    r2 = r + 2 + i
    ws6.cell(row=r2, column=1, value=item).font = BOLD if amount is None else BODY
    if amount is not None:
        ws6.cell(row=r2, column=2, value=amount).font = INPUT_FONT
        ws6.cell(row=r2, column=2).fill = INPUT_YELLOW
    else:
        ws6.cell(row=r2, column=2, value=f'=SUM(B{r+2}:B{r2-1})').font = ALERT_FONT
        ws6.cell(row=r2, column=2).fill = ALERT_RED
    ws6.cell(row=r2, column=2).number_format = '#,##0'
    ws6.cell(row=r2, column=3, value=recipient).font = NOTE
    ws6.cell(row=r2, column=4, value=category).font = NOTE
    for c in range(1, 5): ws6.cell(row=r2, column=c).border = BORDER

r = 33
section_header(ws6, r, "WHAT THE UK GOT BACK", 4)
write_row(ws6, r+1, ['Metric', 'Value', 'Unit', 'Source'], [HEADER]*4, [BLUE]*4)

returns = [
    ('Tax + NI contributed', 5600, 'GBP', 'Payslip data'),
    ('Consumer spending in UK economy', 25000, 'GBP', 'Rent, food, transport, retail'),
    ('NHS services consumed', 150, 'GBP', 'One A&E visit (estimated NHS cost)'),
    ('Prescription cost paid by graduate', 15, 'GBP', 'Out of pocket'),
    ('NHS payments made (NI share + IHS)', 4609, 'GBP', 'NI allocation + Immigration Health Surcharge'),
    ('NHS payment-to-usage ratio', None, 'x', ''),
    ('Net NHS surplus from this graduate', None, 'GBP', ''),
    ('Total tax + fees paid to government', None, 'GBP', ''),
    ('Tuition paid to university', 15000, 'GBP', ''),
    ('TOTAL UK REVENUE FROM THIS GRADUATE', None, 'GBP', ''),
]

for i, (metric, value, unit, source) in enumerate(returns):
    r2 = r + 2 + i
    ws6.cell(row=r2, column=1, value=metric).font = BODY
    if value is not None:
        ws6.cell(row=r2, column=2, value=value).font = INPUT_FONT
        ws6.cell(row=r2, column=2).fill = INPUT_YELLOW
    ws6.cell(row=r2, column=2).number_format = '#,##0'
    ws6.cell(row=r2, column=3, value=unit).font = NOTE
    ws6.cell(row=r2, column=4, value=source).font = NOTE
    for c in range(1, 5): ws6.cell(row=r2, column=c).border = BORDER

# NHS ratio formula
ws6.cell(row=r+7, column=2, value=f'=B{r+6}/B{r+4}').font = OUTPUT_FONT
ws6.cell(row=r+7, column=2).fill = OUTPUT_GREEN
ws6.cell(row=r+7, column=2).number_format = '0.0'
# Net NHS surplus
ws6.cell(row=r+8, column=2, value=f'=B{r+6}-B{r+4}').font = OUTPUT_FONT
ws6.cell(row=r+8, column=2).fill = OUTPUT_GREEN
# Total govt revenue
ws6.cell(row=r+9, column=2, value=f'=B{r+2}+B{r+6}+B21+B22+B23+B24').font = OUTPUT_FONT
ws6.cell(row=r+9, column=2).fill = OUTPUT_GREEN
# Total UK revenue
ws6.cell(row=r+11, column=2, value=f'=B{r+9}+B{r+10}').font = OUTPUT_FONT
ws6.cell(row=r+11, column=2).fill = OUTPUT_GREEN

r = 46
section_header(ws6, r, "EMPLOYABILITY ASSESSMENT", 4)
write_row(ws6, r+1, ['Skill / Qualification', 'Evidence', 'Relevance', 'Verified'], [HEADER]*4, [BLUE]*4)

skills = [
    ('MSc Data Analytics (2:1)', 'Aston University transcript', 'Core qualification', 'Yes'),
    ('Microsoft PL-300 (Power BI)', 'Credential ID: C2A34AA4132D8722', 'Industry certification', 'Online verifiable'),
    ('Google Data Analytics Certificate', 'Coursera — 8 courses', 'Industry training', 'Verifiable'),
    ('Python (pandas, scikit-learn, matplotlib)', '10 GitHub projects with executed notebooks', 'Primary language for data roles', 'Public repos'),
    ('SQL (window functions, CTEs, aggregations)', '14 analytical queries on GitHub', 'Required for all data roles', 'Public repo'),
    ('dbt (Core)', '11 models, 94 automated tests across 2 projects', 'Growing demand in data engineering', 'Public repos'),
    ('Power BI', 'PL-300 certified + dashboard in Apex project', 'Most requested BI tool in UK', 'Verifiable'),
    ('PostgreSQL', '3 projects using PostgreSQL (local + Neon cloud)', 'Standard database', 'Public repos'),
    ('GitHub Actions CI/CD', '3 workflows: lint, test, scheduled ingest', 'DevOps capability', 'Public repo'),
    ('Docker', 'Dockerfile in Apex project', 'Containerisation', 'Public repo'),
    ('AWS (S3, Glue, Athena)', 'Crime Analysis project on AWS', 'Cloud platform', 'Public repo'),
    ('Excel (advanced formulas, charts)', 'Employment analysis + case study workbook', 'Required for analyst roles', 'Public repo'),
    ('Streamlit', 'Live deployed dashboard', 'Data app framework', 'Live URL'),
    ('ERP systems (SI Integreater)', '12 months daily operational use', 'Production environment experience', 'Employer reference'),
    ('Team leadership', 'Managing 11+ operatives at Copernus', 'Soft skill', 'Employer reference'),
]

for i, (skill, evidence, relevance, verified) in enumerate(skills):
    r2 = r + 2 + i
    ws6.cell(row=r2, column=1, value=skill).font = BODY
    ws6.cell(row=r2, column=2, value=evidence).font = NOTE
    ws6.cell(row=r2, column=3, value=relevance).font = NOTE
    ws6.cell(row=r2, column=4, value=verified).font = OUTPUT_FONT if verified in ['Yes', 'Online verifiable', 'Verifiable', 'Live URL'] else BODY
    for c in range(1, 5): ws6.cell(row=r2, column=c).border = BORDER

r = r + 2 + len(skills) + 1
section_header(ws6, r, "PIPELINE OUTCOME", 4)
ws6.cell(row=r+1, column=1, value="Skilled Worker visa salary threshold").font = BODY
ws6.cell(row=r+1, column=2, value="='Model Inputs'!B10").font = OUTPUT_FONT
ws6.cell(row=r+1, column=2).fill = OUTPUT_GREEN
ws6.cell(row=r+1, column=2).number_format = '#,##0'

ws6.cell(row=r+2, column=1, value="Graduate's target salary").font = BODY
ws6.cell(row=r+2, column=2, value=28000).font = INPUT_FONT
ws6.cell(row=r+2, column=2).fill = INPUT_YELLOW
ws6.cell(row=r+2, column=2).number_format = '#,##0'

ws6.cell(row=r+3, column=1, value="Salary gap").font = BOLD
ws6.cell(row=r+3, column=2, value=f'=B{r+1}-B{r+2}').font = ALERT_FONT
ws6.cell(row=r+3, column=2).fill = ALERT_RED
ws6.cell(row=r+3, column=2).number_format = '#,##0'

ws6.cell(row=r+4, column=1, value="Can this graduate qualify for Skilled Worker visa?").font = BOLD
ws6.cell(row=r+4, column=2, value=f'=IF(B{r+3}>0,"NO - Salary gap of "&TEXT(B{r+3},"#,##0")&" prevents sponsorship","YES")').font = ALERT_FONT
ws6.cell(row=r+4, column=2).fill = ALERT_RED

ws6.cell(row=r+6, column=1, value="CONCLUSION").font = H2
ws6.cell(row=r+7, column=1, value="This graduate invested 51,523 into the UK economy. They hold a postgraduate degree,").font = BODY
ws6.cell(row=r+8, column=1, value="an industry certification, 15 verified technical skills, and 10 portfolio projects.").font = BODY
ws6.cell(row=r+9, column=1, value="The salary threshold prevents them from staying. The pipeline loses this talent.").font = BODY
ws6.cell(row=r+10, column=1, value="This is one case. Multiply by 100,000 graduates per year.").font = BOLD

output = 'UK_Graduate_Talent_Pipeline_Analysis.xlsx'
wb.save(output)
print(f'Saved: {output}')
print(f'Sheets: {[ws.title for ws in wb.worksheets]}')
